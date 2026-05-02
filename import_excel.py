"""
Importa los datos existentes del Excel a la base de datos.
Ejecutar una sola vez: python3 import_excel.py
"""
import os
import sys
import sqlite3
import openpyxl

EXCEL_PATH = os.path.expanduser(
    '~/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/'
    '5447963F-CEAB-4BAA-9E71-6E5B9EF1F4FB/Golf_Guadalmina_2026_Ranking_COMPLETO.xlsx'
)

sys.path.insert(0, os.path.dirname(__file__))
import database as db

MESES_NUM = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
    'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
    'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
}


def run():
    db.init_db()

    conn = db.get_conn()

    # Comprobamos si ya hay datos
    if conn.execute("SELECT COUNT(*) FROM jugadores").fetchone()[0] > 0:
        print("La base de datos ya tiene datos. Limpiando para reimportar...")
        conn.execute("DELETE FROM resultados")
        conn.execute("DELETE FROM partidas")
        conn.execute("DELETE FROM jugadores")
        conn.commit()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Jugadores ──────────────────────────────────────────────────────────────
    ws_j = wb['Jugadores']
    jugador_map = {}  # nombre → id
    for row in ws_j.iter_rows(min_row=4, max_row=25, values_only=True):
        if row[1] is not None and row[2] is not None:
            nombre = str(row[1]).strip()
            hcp    = float(row[2])
            conn.execute("INSERT INTO jugadores (nombre, hcp) VALUES (?,?)", (nombre, hcp))
    conn.commit()

    for j in conn.execute("SELECT id, nombre FROM jugadores").fetchall():
        jugador_map[j['nombre']] = j['id']
    print(f"Importados {len(jugador_map)} jugadores.")

    # ── Resultados ─────────────────────────────────────────────────────────────
    ws_r = wb['Resultados']
    rows = list(ws_r.iter_rows(min_row=5, max_row=305, values_only=True))

    # Agrupar por (dia, mes, año) → partida
    from collections import defaultdict
    partidas_data = defaultdict(list)

    semana_counter = {}   # (dia, mes, ano) → semana

    for row in rows:
        sem, dia, mes, ano, campo, hoyos, jugador, hcp, stableford, pos, pts = row[:11]
        if not jugador or not stableford or not mes:
            continue
        jugador = str(jugador).strip()
        key = (dia, mes, ano)
        partidas_data[key].append({
            'semana': sem,
            'campo': campo or 'Guadalmina Norte',
            'hoyos': hoyos or '9 hoyos',
            'jugador': jugador,
            'hcp': float(hcp) if hcp else None,
            'stableford': int(stableford),
        })

    partidas_insertadas = 0
    for (dia, mes, ano), filas in sorted(partidas_data.items(),
                                          key=lambda x: (x[0][2], MESES_NUM.get(x[0][1], 0), x[0][0])):
        campo = filas[0]['campo']
        hoyos = filas[0]['hoyos']
        semana = next((f['semana'] for f in filas if f['semana']), None)
        mes_num = MESES_NUM.get(mes, 1)
        fecha = f"{ano}-{mes_num:02d}-{int(dia):02d}"

        cur = conn.execute(
            "INSERT INTO partidas (semana, fecha, campo, hoyos, es_major) VALUES (?,?,?,?,0)",
            (semana, fecha, campo, hoyos)
        )
        partida_id = cur.lastrowid

        for f in filas:
            jid = jugador_map.get(f['jugador'])
            if not jid:
                print(f"  AVISO: jugador '{f['jugador']}' no encontrado, ignorando.")
                continue
            hcp_val = f['hcp']
            if hcp_val is None:
                hcp_val = conn.execute("SELECT hcp FROM jugadores WHERE id=?", (jid,)).fetchone()['hcp']
            conn.execute(
                "INSERT INTO resultados (partida_id, jugador_id, hcp, stableford) VALUES (?,?,?,?)",
                (partida_id, jid, hcp_val, f['stableford'])
            )

        conn.commit()
        db.calcular_posicion_y_puntos(partida_id)
        partidas_insertadas += 1
        print(f"  Partida {fecha} ({hoyos}) — {len(filas)} jugadores")

    print(f"\nImportación completada: {partidas_insertadas} partidas, {len(jugador_map)} jugadores.")


if __name__ == '__main__':
    run()
